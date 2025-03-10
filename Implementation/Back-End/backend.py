# Here is a list of backend functions that will be implemented according to our API design.
# These functions will be exclusively accessed from inside this file, or from api.py.
# The database instance will only be accessed from this file.
#####################################################################################################################################
# The below database connection will be the only one used in this file.
# If you need to reset the cursor, you can do so by calling cursor.close() and then cursor = database.cursor()
# If you need to reset the database, you can do so by calling database.close() and then database = sqlite3.connect("database.db")
# Only do this if necessary and do it in your own functions.
import sqlite3
import pandas as pd
import xlsxwriter
database = sqlite3.connect("database.db")
cursor = database.cursor()

# Assigned to: Charlie
# This function will load a XLSX file from a given location and return it as a dictionary
import openpyxl
from io import BytesIO
def load_xlsx(fileLocation: str = "", fileBytes: bytes = None) -> dict:
    # Load the workbook
    if fileBytes is None:
        workbook = openpyxl.load_workbook(fileLocation)
    else:
        workbook = openpyxl.load_workbook(BytesIO(fileBytes))
    
    # Initialize the dictionary to store the data
    data_dict = {}
    
    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Initialize a list to store rows for the current sheet
        rows = []
        
        # Iterate through each row in the sheet
        for row in sheet.iter_rows(values_only=True):
            # Append the row (as a tuple) to the rows list
            rows.append(row)
        
        # Add the rows list to the dictionary with the sheet name as the key
        data_dict[sheet_name] = rows
    
    return data_dict

def DLNewspaperReferences2025():
    cursor.execute("SELECT * FROM NewspaperReferences2025")
    Data = cursor.fetchall()
    Headers = [description[0] for description in cursor.description]
    HeadersWithData = {}
    for num in range(len(Headers)-1):
        HeadersWithData = {**HeadersWithData, Headers[num+1]: [d[num+1] for d in Data]}
    df = pd.DataFrame(HeadersWithData)
    writer = pd.ExcelWriter("NewspaperReferences2025.xlsx",engine="xlsxwriter")
    df.to_excel(writer,sheet_name="sheet1",index=False)
    workbook = writer.book
    worksheet = writer.sheets["sheet1"]
    worksheet.autofit()
    workbook.close()

    # RollOfHonour
def DLRollOfHonour():
    cursor.execute("SELECT * FROM RollOfHonour")
    Data = cursor.fetchall()
    Headers = [description[0] for description in cursor.description]
    HeadersWithData = {}
    for num in range(len(Headers) - 1):
        HeadersWithData = {**HeadersWithData, Headers[num + 1]: [d[num + 1] for d in Data]}
    df = pd.DataFrame(HeadersWithData)
    writer = pd.ExcelWriter("RollOfHonour.xlsx", engine="xlsxwriter")
    df.to_excel(writer, sheet_name="sheet1", index=False)
    workbook = writer.book
    worksheet = writer.sheets["sheet1"]
    worksheet.autofit()
    workbook.close()

    # BiographySpreadsheet
def DLBiographySpreadsheet():
    cursor.execute("SELECT * FROM BiographySpreadsheet")
    Data = cursor.fetchall()
    Headers = [description[0] for description in cursor.description]
    HeadersWithData = {}
    for num in range(len(Headers) - 1):
        HeadersWithData = {**HeadersWithData, Headers[num + 1]: [d[num + 1] for d in Data]}
    df = pd.DataFrame(HeadersWithData)
    writer = pd.ExcelWriter("BiographySpreadsheet.xlsx", engine="xlsxwriter")
    df.to_excel(writer, sheet_name="sheet1", index=False)
    workbook = writer.book
    worksheet = writer.sheets["sheet1"]
    worksheet.autofit()
    workbook.close()

    # BradfordMemorials
def DLBradfordMemorials():
    cursor.execute("SELECT * FROM BradfordMemorials")
    Data = cursor.fetchall()
    Headers = [description[0] for description in cursor.description]
    HeadersWithData = {}
    for num in range(len(Headers) - 1):
        HeadersWithData = {**HeadersWithData, Headers[num + 1]: [d[num + 1] for d in Data]}
    df = pd.DataFrame(HeadersWithData)
    writer = pd.ExcelWriter("BradfordMemorials.xlsx", engine="xlsxwriter")
    df.to_excel(writer, sheet_name="sheet1", index=False)
    workbook = writer.book
    worksheet = writer.sheets["sheet1"]
    worksheet.autofit()
    workbook.close()

    # BuriedInBradford
def DLBuriedInBradford():
    cursor.execute("SELECT * FROM BuriedInBradford")
    Data = cursor.fetchall()
    Headers = [description[0] for description in cursor.description]
    HeadersWithData = {}
    for num in range(len(Headers) - 1):
        HeadersWithData = {**HeadersWithData, Headers[num + 1]: [d[num + 1] for d in Data]}
    df = pd.DataFrame(HeadersWithData)
    writer = pd.ExcelWriter("BuriedInBradford.xlsx", engine="xlsxwriter")
    df.to_excel(writer, sheet_name="sheet1", index=False)
    workbook = writer.book
    worksheet = writer.sheets["sheet1"]
    worksheet.autofit()
    workbook.close()


# Assigned to: Edward
#This function should search for a name
def forename_S(desired_DB, forename_S):
    sql_fixed = f"SELECT * FROM `{desired_DB}` WHERE forename LIKE '%{forename_S}%'"
    cursor.execute(sql_fixed)
    return cursor.fetchall()

def surname_S(desired_DB, surname_S):
    sql_fixed = f"SELECT * FROM `{desired_DB}` WHERE surname LIKE '%{surname_S}%'"
    cursor.execute(sql_fixed)
    return cursor.fetchall()

def regiment_S(desired_DB, regiment_S):
    sql_fixed = f"SELECT * FROM `{desired_DB}` WHERE regiment LIKE '%{regiment_S}%'"
    cursor.execute(sql_fixed)
    return cursor.fetchall()

def forename_surname_S(desired_DB, forename_S, surname_S):
    sql_fixed = f"SELECT * FROM `{desired_DB}` WHERE forename LIKE '%{forename_S}%' AND surname LIKE '%{surname_S}%'"
    cursor.execute(sql_fixed)
    return cursor.fetchall()

def forename_regiment_S(desired_DB, forename_S, regiment_S):
    sql_fixed = f"SELECT * FROM `{desired_DB}` WHERE forename LIKE '%{forename_S}%' AND regiment LIKE '%{regiment_S}%'"
    cursor.execute(sql_fixed)
    return cursor.fetchall()

def surname_regiment_S(desired_DB, surname_S, regiment_S):
    sql_fixed = f"SELECT * FROM `{desired_DB}` WHERE surname LIKE '%{surname_S}%' AND regiment LIKE '%{regiment_S}%'"
    cursor.execute(sql_fixed)
    return cursor.fetchall()

def forename_surname_regiment_S(desired_DB, forename_S, surname_S, regiment_S):
    sql_fixed = f"SELECT * FROM `{desired_DB}` WHERE forename LIKE '%{forename_S}%' AND surname LIKE '%{surname_S}%' AND regiment LIKE '%{regiment_S}%'"
    cursor.execute(sql_fixed)
    return cursor.fetchall()


# Assigned to: Hope
# This function will instantiate the databases as well as their columns for the next function, insert_to_sql()
def create_database ():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS biographyspreadsheet (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            surname VARCHAR(80),
            forename VARCHAR(80),
            regiment VARCHAR(70),
            service_number VARCHAR(40),
            biography_attachment VARCHAR(300)
        );
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bradfordmemorials (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            surname VARCHAR(80),
            forename VARCHAR(80) ,
            regiment VARCHAR(70),
            unit VARCHAR(40),
            memorial VARCHAR(150),
            memorial_location VARCHAR(150),
            memorial_info VARCHAR(150),
            memorial_postcode VARCHAR(32),
            district VARCHAR(150),
            photo_available BOOLEAN
        );
     """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS buriedinbradford (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            surname VARCHAR(80),
            forename VARCHAR(80),
            age INT(3),
            medals LONGTEXT,
            date_of_birth DATE,
            rank VARCHAR(40),
            unit VARCHAR(40),
            cemetery VARCHAR(150),
            grave_ref VARCHAR(40),
            info LONGTEXT
        );
     """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS memorialnames (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            surname VARCHAR(80),
            forename VARCHAR(80),
            memorial VARCHAR(150)
         );
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS newspaperreferences2025 (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            surname VARCHAR(80),
            forename VARCHAR(80),
            rank VARCHAR(40),
            address VARCHAR(150),
            regiment VARCHAR(70),
            unit VARCHAR(40),
            article_comment VARCHAR(300),
            newspaper_name VARCHAR(150),
            newspaper_date DATE,
            page_col VARCHAR(10),
            photo_incl BOOLEAN
        );
     """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS rollofhonour (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            surname VARCHAR(80),
            forename VARCHAR(80),
            address VARCHAR(150),
            electoral_ward VARCHAR(35),
            town VARCHAR(35),
            rank VARCHAR(40),
            regiment VARCHAR(70),
            unit VARCHAR(40),
            company VARCHAR(40),
            age INT(3),
            service_no VARCHAR(40),
            other_regiment VARCHAR(70),
            other_service_no VARCHAR(40),
            medals LONGTEXT,
            enlisted_date DATE,
            discharged_date DATE,
            death_date DATE,
            misc_info_nroh VARCHAR(200),
            cemetery_memorial VARCHAR(150),
            cemetery_memorial_ref VARCHAR(40),
            cemetery_memorial_country VARCHAR(56),
            additional_cwgc_info VARCHAR(300)
        );
     """)

    return True

def drop_databases():
    # Drop all tables
    cursor = database.cursor()
    cursor.execute("DROP TABLE IF EXISTS BradfordMemorials")
    cursor.execute("DROP TABLE IF EXISTS NewspaperReferences2025")
    cursor.execute("DROP TABLE IF EXISTS BiographySpreadsheet")
    cursor.execute("DROP TABLE IF EXISTS RollOfHonour")
    cursor.execute("DROP TABLE IF EXISTS BuriedInBradford")

# Assigned to: Charlie
# This function will call the load_csv() on all of the csv files in the directory
# It will then convert their returned values into the created SQL databases in create_database()
import datetime
def insert_to_sql(sheetData: dict) -> bool:
    # Detect which database the data belongs in
    desired_table = detect_table(sheetData)

    insert_statements = {
        "BradfordMemorials": "INSERT INTO BradfordMemorials (surname, forename, regiment, unit, memorial, memorial_location, memorial_info, memorial_postcode, district, photo_available) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);",
        "NewspaperReferences2025": "INSERT INTO NewspaperReferences2025 (surname, forename, rank, address, regiment, unit, article_comment, newspaper_name, newspaper_date, page_col, photo_incl) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);",
        "BiographySpreadsheet": "INSERT INTO BiographySpreadsheet (surname, forename, regiment, service_number, biography_attachment) VALUES (?, ?, ?, ?, ?);",
        "RollOfHonour": "INSERT INTO RollOfHonour (surname, forename, address, electoral_ward, town, rank, regiment, unit, company, age, service_no, other_regiment, other_service_no, medals, enlisted_date, discharged_date, death_date, misc_info_nroh, cemetery_memorial, cemetery_memorial_ref, cemetery_memorial_country, additional_cwgc_info) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);",
        "BuriedInBradford": "INSERT INTO BuriedInBradford (surname, forename, age, medals, date_of_birth, rank, unit, cemetery, grave_ref, info) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);"
    }

    create_database()

    # Insert the data into the database
    if desired_table == "BradfordMemorials":
        # init variables for spreadsheet filtering
        iterator = 0
        filter_data = {
                        "Names on Bradford Memorials": {"max_rows": 2, "skip_rows": 4}
                      }
        sheet_names = ["Names on Bradford Memorials"]

        for sheet_name in sheet_names:
            # count all the rows without None values
            row_count = sheetData[sheet_name].__len__()
            row_count -= filter_data[sheet_name]["skip_rows"]

            max_rows_iter = row_count
            for row in sheetData[sheet_name]:
                if max_rows_iter == 0:
                    break
                if iterator < filter_data[sheet_name]["skip_rows"]:
                    iterator += 1
                    continue

                surname = row[0]
                # format rank by converting " and ' to web safe characters
                if surname is not None:
                    surname = surname.replace('"', "&quot;").replace("'", "&apos;")
                forename = row[1]
                # format rank by converting " and ' to web safe characters
                if forename is not None:
                    forename = forename.replace('"', "&quot;").replace("'", "&apos;")
                regiment = row[2]
                # format rank by converting " and ' to web safe characters
                if regiment is not None:
                    regiment = regiment.replace('"', "&quot;").replace("'", "&apos;")
                unit = row[3]
                # format rank by converting " and ' to web safe characters
                if unit is not None:
                    unit = unit.replace('"', "&quot;").replace("'", "&apos;")
                memorial = row[4]
                # format rank by converting " and ' to web safe characters
                if memorial is not None:
                    memorial = memorial.replace('"', "&quot;").replace("'", "&apos;")
                memorial_location = row[5]
                # format rank by converting " and ' to web safe characters
                if memorial_location is not None:
                    memorial_location = memorial_location.replace('"', "&quot;").replace("'", "&apos;")
                memorial_info = row[6]
                # format rank by converting " and ' to web safe characters
                if memorial_info is not None:
                    memorial_info = memorial_info.replace('"', "&quot;").replace("'", "&apos;")
                memorial_postcode = row[7]
                # format rank by converting " and ' to web safe characters
                if memorial_postcode is not None:
                    memorial_postcode = memorial_postcode.replace('"', "&quot;").replace("'", "&apos;")
                district = row[8]
                # format rank by converting " and ' to web safe characters
                if district is not None:
                    district = district.replace('"', "&quot;").replace("'", "&apos;")
                photo_available = row[9]
                # format rank by converting " and ' to web safe characters
                if photo_available is not None:
                    photo_available = photo_available.replace('"', "&quot;").replace("'", "&apos;")
                
                cursor = database.cursor()

                completed_statement = ""+insert_statements[desired_table]
                replaced = completed_statement.replace("?", "'{}'").format(surname, forename, regiment, unit, memorial, memorial_location, memorial_info, memorial_postcode, district, photo_available)

                cursor.execute(replaced)

                max_rows_iter -= 1
            database.commit()

        print(f"Inserting data into {desired_table}...")
    elif desired_table == "NewspaperReferences2025":
        # init variables for spreadsheet filtering
        iterator = 0
        filter_data = {
                        "Sheet1": {"max_rows": 2, "skip_rows": 1}
                      }
        sheet_names = ["Sheet1"]

        for sheet_name in sheet_names:
            # count all the rows without None values
            row_count = sheetData[sheet_name].__len__()
            row_count -= filter_data[sheet_name]["skip_rows"]

            max_rows_iter = row_count
            for row in sheetData[sheet_name]:
                if max_rows_iter == 0:
                    break
                if iterator < filter_data[sheet_name]["skip_rows"]:
                    iterator += 1
                    continue

                surname = row[0]
                # format rank by converting " and ' to web safe characters
                if surname is not None:
                    surname = surname.replace('"', "&quot;").replace("'", "&apos;")
                forename = row[1]
                # format rank by converting " and ' to web safe characters
                if forename is not None:
                    forename = forename.replace('"', "&quot;").replace("'", "&apos;")
                rank = row[2]
                # format rank by converting " and ' to web safe characters
                if rank is not None:
                    rank = rank.replace('"', "&quot;").replace("'", "&apos;")
                address = row[3]
                # format rank by converting " and ' to web safe characters
                if address is not None:
                    address = address.replace('"', "&quot;").replace("'", "&apos;")
                regiment = row[4]
                # format rank by converting " and ' to web safe characters
                if regiment is not None:
                    regiment = regiment.replace('"', "&quot;").replace("'", "&apos;")
                unit = row[5]
                # format rank by converting " and ' to web safe characters
                if unit is not None:
                    unit = unit.replace('"', "&quot;").replace("'", "&apos;")
                article_comment = row[6]
                # format rank by converting " and ' to web safe characters
                if article_comment is not None:
                    article_comment = article_comment.replace('"', "&quot;").replace("'", "&apos;")
                newspaper_name = row[7]
                # format rank by converting " and ' to web safe characters
                if newspaper_name is not None:
                    newspaper_name = newspaper_name.replace('"', "&quot;").replace("'", "&apos;")
                newspaper_date = row[8]
                page_col = row[9]
                # format rank by converting " and ' to web safe characters
                if page_col is not None:
                    page_col = page_col.replace('"', "&quot;").replace("'", "&apos;")
                photo_incl = row[10]
                # format rank by converting " and ' to web safe characters
                if photo_incl is not None:
                    photo_incl = photo_incl.replace('"', "&quot;").replace("'", "&apos;")
                
                cursor = database.cursor()

                completed_statement = ""+insert_statements[desired_table]
                replaced = completed_statement.replace("?", "'{}'").format(surname, forename, rank, address, regiment, unit, article_comment, newspaper_name, newspaper_date, page_col, photo_incl)

                cursor.execute(replaced)

                max_rows_iter -= 1
            database.commit()

        print(f"Inserting data into {desired_table}...")
    elif desired_table == "BiographySpreadsheet":
        # init variables for spreadsheet filtering
        iterator = 0
        filter_data = {
                        "Sheet1": {"max_rows": 2, "skip_rows": 3}
                      }
        sheet_names = ["Sheet1"]

        for sheet_name in sheet_names:
            # count all the rows without None values
            row_count = sheetData[sheet_name].__len__()
            row_count -= filter_data[sheet_name]["skip_rows"]

            max_rows_iter = row_count
            for row in sheetData[sheet_name]:
                if max_rows_iter == 0:
                    break
                if iterator < filter_data[sheet_name]["skip_rows"]:
                    iterator += 1
                    continue

                surname = row[1]
                # format rank by converting " and ' to web safe characters
                if surname is not None:
                    surname = surname.replace('"', "&quot;").replace("'", "&apos;")
                forename = row[2]
                # format rank by converting " and ' to web safe characters
                if forename is not None:
                    forename = forename.replace('"', "&quot;").replace("'", "&apos;")
                regiment = row[3]
                # format rank by converting " and ' to web safe characters
                if regiment is not None:
                    regiment = regiment.replace('"', "&quot;").replace("'", "&apos;")
                service_number = row[4]
                # format rank by converting " and ' to web safe characters
                if service_number is not None:
                    service_number = str(service_number).replace('"', "&quot;").replace("'", "&apos;")
                biography_attachment = row[5]
                # format rank by converting " and ' to web safe characters
                if biography_attachment is not None:
                    biography_attachment = biography_attachment.replace('"', "&quot;").replace("'", "&apos;")
                
                cursor = database.cursor()

                completed_statement = ""+insert_statements[desired_table]
                replaced = completed_statement.replace("?", "'{}'").format(surname, forename, regiment, service_number, biography_attachment)

                cursor.execute(replaced)

                max_rows_iter -= 1
            database.commit()

        print(f"Inserting data into {desired_table}...")
    elif desired_table == "RollOfHonour":
        iterator = 0
        filter_data = {
                        "Sheet1": {"max_rows": 2, "skip_rows": 3}
                      }
        sheet_names = ["Sheet1"]

        for sheet_name in sheet_names:
            # count all the rows without None values
            row_count = sheetData[sheet_name].__len__()
            row_count -= filter_data[sheet_name]["skip_rows"]

            max_rows_iter = row_count
            for row in sheetData[sheet_name]:
                if max_rows_iter == 0:
                    break
                if iterator < filter_data[sheet_name]["skip_rows"]:
                    iterator += 1
                    continue
                
                surname = row[0]
                # format rank by converting " and ' to web safe characters
                if surname is not None:
                    surname = surname.replace('"', "&quot;").replace("'", "&apos;")
                forename = row[1]
                # format rank by converting " and ' to web safe characters
                if forename is not None:
                    forename = forename.replace('"', "&quot;").replace("'", "&apos;")
                address = row[2]
                # format rank by converting " and ' to web safe characters
                if address is not None:
                    address = address.replace('"', "&quot;").replace("'", "&apos;")
                electoral_ward = row[3]
                # format rank by converting " and ' to web safe characters
                if electoral_ward is not None:
                    electoral_ward = electoral_ward.replace('"', "&quot;").replace("'", "&apos;")
                town = row[4]
                # format rank by converting " and ' to web safe characters
                if town is not None:
                    town = town.replace('"', "&quot;").replace("'", "&apos;")
                rank = row[5]
                # format rank by converting " and ' to web safe characters
                if rank is not None:
                    rank = rank.replace('"', "&quot;").replace("'", "&apos;")
                regiment = row[6]
                # format rank by converting " and ' to web safe characters
                if regiment is not None:
                    regiment = regiment.replace('"', "&quot;").replace("'", "&apos;")
                unit = row[7]
                # format rank by converting " and ' to web safe characters
                if unit is not None:
                    unit = str(unit).replace('"', "&quot;").replace("'", "&apos;")
                company = row[8]
                # format rank by converting " and ' to web safe characters
                if company is not None:
                    company = company.replace('"', "&quot;").replace("'", "&apos;")
                age = row[9]
                service_no = row[10]
                if service_no is not None:
                    service_no = str(service_no).replace('"', "&quot;").replace("'", "&apos;")
                other_regiment = row[11]
                # format rank by converting " and ' to web safe characters
                if other_regiment is not None:
                    other_regiment = other_regiment.replace('"', "&quot;").replace("'", "&apos;")
                other_service_no = row[13]
                if other_service_no is not None:
                    other_service_no = str(other_service_no).replace('"', "&quot;").replace("'", "&apos;")
                medals = row[14]
                # format rank by converting " and ' to web safe characters
                if medals is not None:
                    medals = medals.replace('"', "&quot;").replace("'", "&apos;")
                if isinstance(row[15], datetime.datetime):
                    enlisted_date = row[15].strftime("%Y-%m-%d") if row[15] is not None else None
                else:
                    enlisted_date = None
                if isinstance(row[16], datetime.datetime):
                    discharge_date = row[16].strftime("%Y-%m-%d") if row[16] is not None else None
                else:
                    discharge_date = None
                if isinstance(row[17], datetime.datetime):
                    death_date = row[17].strftime("%Y-%m-%d") if row[17] is not None else None
                else:
                    death_date = None
                misc_info_nroh = row[18]
                # format rank by converting " and ' to web safe characters
                if misc_info_nroh is not None:
                    misc_info_nroh = misc_info_nroh.replace('"', "&quot;").replace("'", "&apos;")
                cemetery_memorial = row[19]
                # format rank by converting " and ' to web safe characters
                if cemetery_memorial is not None:
                    cemetery_memorial = cemetery_memorial.replace('"', "&quot;").replace("'", "&apos;")
                cemetery_memorial_ref = row[20]
                # format rank by converting " and ' to web safe characters
                if not isinstance(cemetery_memorial_ref, int):
                    if cemetery_memorial_ref is not None:
                        cemetery_memorial_ref = cemetery_memorial_ref.replace('"', "&quot;").replace("'", "&apos;")
                cemetery_memorial_country = row[21]
                # format rank by converting " and ' to web safe characters
                if cemetery_memorial_country is not None:
                    cemetery_memorial_country = cemetery_memorial_country.replace('"', "&quot;").replace("'", "&apos;")
                additional_cwgc_info = row[22]
                # format rank by converting " and ' to web safe characters
                if additional_cwgc_info is not None:
                    additional_cwgc_info = additional_cwgc_info.replace('"', "&quot;").replace("'", "&apos;")

                # Replace all "Not recorded" string values with None
                if surname == "Not recorded":
                    surname = None
                if forename == "Not recorded":
                    forename = None
                if address == "Not recorded":
                    address = None
                if electoral_ward == "Not recorded":
                    electoral_ward = None
                if town == "Not recorded":
                    town = None
                if rank == "Not recorded":
                    rank = None
                if regiment == "Not recorded":
                    regiment = None
                if unit == "Not recorded":
                    unit = None
                if company == "Not recorded":
                    company = None
                if age == "Not recorded":
                    age = None
                if service_no == "Not recorded":
                    service_no = None
                if other_regiment == "Not recorded":
                    other_regiment = None
                if other_service_no == "Not recorded":
                    other_service_no = None
                if medals == "Not recorded":
                    medals = None
                if enlisted_date == "Not recorded":
                    enlisted_date = None
                if discharge_date == "Not recorded":
                    discharge_date = None
                if death_date == "Not recorded":
                    death_date = None
                if misc_info_nroh == "Not recorded":
                    misc_info_nroh = None
                if cemetery_memorial == "Not recorded":
                    cemetery_memorial = None
                if cemetery_memorial_ref == "Not recorded":
                    cemetery_memorial_ref = None
                if cemetery_memorial_country == "Not recorded":
                    cemetery_memorial_country = None
                if additional_cwgc_info == "Not recorded":
                    additional_cwgc_info = None
                
                cursor = database.cursor()

                completed_statement = ""+insert_statements[desired_table]
                replaced = completed_statement.replace("?", "'{}'").format(surname, forename, address, electoral_ward, town, rank, regiment, unit, company, age, service_no, other_regiment, other_service_no, medals, enlisted_date, discharge_date, death_date, misc_info_nroh, cemetery_memorial, cemetery_memorial_ref, cemetery_memorial_country, additional_cwgc_info)

                cursor.execute(replaced)

                max_rows_iter -= 1
            database.commit()
        print(f"Inserting data into {desired_table}...")
    elif desired_table == "BuriedInBradford":
        # init variables for spreadsheet filtering
        iterator = 0
        filter_data = {
                        "WW1 Burials in Bradford": {"max_rows": 2, "skip_rows": 2}
                      }
        sheet_names = ["WW1 Burials in Bradford"]

        for sheet_name in sheet_names:
            # count all the rows without None values
            row_count = 0
            for row in sheetData[sheet_name]:
                if None not in row:
                    row_count += 1
            row_count -= 3

            max_rows_iter = row_count
            for row in sheetData[sheet_name]:
                if max_rows_iter == 0:
                    break
                if iterator < filter_data[sheet_name]["skip_rows"]:
                    iterator += 1
                    continue
                
                surname = row[0]
                # format rank by converting " and ' to web safe characters
                surname = surname.replace('"', "&quot;").replace("'", "&apos;")
                forename = row[1]
                # format rank by converting " and ' to web safe characters
                forename = forename.replace('"', "&quot;").replace("'", "&apos;")
                age = row[2]
                medals = row[3]
                if isinstance(row[4], datetime.datetime):
                    date_of_birth = row[4].strftime("%Y-%m-%d") if row[4] is not None else None
                else:
                    date_of_birth = None
                rank = row[5]
                # format rank by converting " and ' to web safe characters
                rank = rank.replace('"', "&quot;").replace("'", "&apos;")
                unit = row[6]
                cemetery = row[9]
                grave_ref = row[10]
                info = row[11]
                # format info by converting " and ' to web safe characters
                info = info.replace('"', "&quot;").replace("'", "&apos;")
                
                cursor = database.cursor()

                completed_statement = ""+insert_statements[desired_table]
                replaced = completed_statement.replace("?", "'{}'").format(surname, forename, age, medals, date_of_birth, rank, unit, cemetery, grave_ref, info)

                cursor.execute(replaced)

                max_rows_iter -= 1
            database.commit()

        print(f"Inserting data into {desired_table}...")

    return True

def detect_table(sheetData: dict) -> str:
    first_layer_keys = list(sheetData.keys())
    table_name_detection = "Unknown"
    fk = first_layer_keys[0]
    if fk == "Analytics":
        table_name_detection = "BuriedInBradford"
    elif fk == "Bradford CWGC":
        table_name_detection = "BradfordMemorials"
    elif fk == "Sheet1":
        iter = 0
        for sheet in sheetData:
            if iter == 0:
                first_row = sheetData[sheet][0]
                no_of_values = len(first_row)

                if no_of_values == 6:
                    table_name_detection = "BiographySpreadsheet"
                elif no_of_values == 23:
                    table_name_detection = "RollOfHonour"
                elif no_of_values == 11:
                    table_name_detection = "NewspaperReferences2025"
            iter += 1

    return table_name_detection

# Assigned to: ???
# This function will add a singular row to the SQL database
def singular_add_sql(dataObject: dict, databaseName: str):
    return True

def is_first_run() -> bool:
    # Create a file to store if the first run has been ran after returning if it is first run
    try:
        with open("firstrun.flag", "r") as file:
            return False
    except FileNotFoundError:
        with open("firstrun.flag", "w") as file:
            file.write("First run has been executed.")
        return True

# Do not edit any code below this point.
# Only edit code you have been assigned in this project.
if __name__ == "__main__":
    if is_first_run(): #replace with first_run check
        drop_databases()
        insert_to_sql(load_xlsx("Tests Data/Those buried in Bradford.xlsx"))
        insert_to_sql(load_xlsx("Tests Data/Bradford and surrounding townships Great War Roll of Honour 2025.xlsx"))
        insert_to_sql(load_xlsx("Tests Data/Bradford Memorials.xlsx"))
        insert_to_sql(load_xlsx("Tests Data/Newspaper references 2025.xlsx"))
        insert_to_sql(load_xlsx("Tests Data/Biography spreadsheet.xlsx"))
    exit()
